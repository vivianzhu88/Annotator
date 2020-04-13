import urllib.request, urllib.error, urllib.parse
import json
import os
import openpyxl
import time
from pprint import pprint
from socket import error as SocketError
import errno
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import docx

REST_URL = "http://data.bioontology.org"
ONT = "&ontologies=RADLEX"
API_KEY = ""

def toSpreadsheet(filesList):
#put filename and RIDs into to Excel spreadsheet
    #RIDs
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row = 1
    for f in filesList:
        if f.getRIDs():
            sheet.cell(row=row, column=1, value=f.getName())
            sheet.cell(row=row, column=2, value=', '.join(f.getRIDs()))
            row += 1

    workbook.save(filename="fileRIDs.xlsx")
    
    #Rterms
    workbook2 = openpyxl.Workbook()
    sheet = workbook2.active
    row = 1
    for f in filesList:
        if f.getRIDs():
            sheet.cell(row=row, column=1, value=f.getName())
            sheet.cell(row=row, column=2, value=', '.join(f.getRterms()))
            row += 1

    workbook2.save(filename="fileRterms.xlsx")

class File():
    def __init__(self, filename):
    #initialize each file
        self.filename = filename
        self.RIDs = []
        self.Rterms = []
        self.text = ""
    
    def getName(self):
    #return filename
        return self.filename[64:]
    
    def getRIDs(self):
    #return RIDs list
        return self.RIDs
    
    def getRterms(self):
    #return Rterms list
        return self.Rterms
    
    def openFile(self):
    #open file and retrieve text
    #has .ore .txt .lsx .ocx .csv .xls .son .xml files
        try:
            with open(self.filename, 'r') as file:
                self.text = file.read()
            return
        except UnicodeDecodeError:
            pass
        
        try:
            df = pd.read_excel(self.filename)
            self.text = df.to_string()
            return
        except:
            pass

        try:
            doc = docx.Document(self.filename)
            fullText = []
            for para in doc.paragraphs:
                txt = para.text.encode('ascii', 'ignore')
                fullText.append(txt)
            self.text = b'\n'.join(fullText)
        except:
            print(self.filename)
            
    def get_json(self, url):
    #get json from annotator
        opener = urllib.request.build_opener()
        opener.addheaders = [('Authorization', 'apikey token=' + API_KEY)]
        return json.loads(opener.open(url).read())
        
    def split_text(self, number):
    #split text into smaller pieces for Annotator to handle
        if len(self.text) > 1:
            chunks, chunk_size = len(self.text), len(self.text)//number
            return [self.text[i:i+chunk_size] for i in range(0, chunks, chunk_size)]
        return [self.text]
        
    def mapRIDs(self, annotations, get_class=True):
    #take annotations and map to Excel spreadsheet of labels and corresponding RIDs
        #iterate through annotations
        for result in annotations:
            class_details = result["annotatedClass"]
            if get_class:
                try:
                    class_details = self.get_json(result["annotatedClass"]["links"]["self"])
                except urllib.error.HTTPError:
                    print(f"Error retrieving {result['annotatedClass']['@id']}")
                    continue
            
            #get each RIDs + remove duplicates
            id = class_details["@id"]
            rid = id[22:]
            
            if rid not in self.RIDs:
                self.RIDs.append(id[22:])
                
            #get each Rterms + remove duplicates
            rterm = class_details["prefLabel"]
            
            if rterm not in self.Rterms:
                self.Rterms.append(rterm)
    
    def getContents(self):
    #runs all the methods needed to parse files and get annotations
        self.openFile()
        
        i = 1
        try:
            texts = self.split_text(i)
            for t in texts:
                try:
                    annotations = self.get_json(REST_URL + "/annotator?text=" + urllib.parse.quote(t) + ONT)
                    self.mapRIDs(annotations)
                except SocketError as e:
                    pass
        except urllib.error.HTTPError:
            i += 1
            
#put all of file paths in Chest_and_Lung_Collections directory into a list
filesList = []

dir_path = "/Users/vivianzhu/Documents/Annotator/Chest_and_Lung_Collections/"
for dp,_,filenames in os.walk(dir_path):
   for f in filenames:
       if f != ".DS_Store":
           f = File(os.path.abspath(os.path.join(dp, f)))
           filesList.append(f)

for f in filesList:
    f.getContents()
    
toSpreadsheet(filesList)
    
    
